"""月度組合報告 — 由 GitHub Actions 每月 1 日 09:00 TPE 自動執行。

主要功能:
  1. 載入持股 xlsx 取得目前組合狀態 (沿用 update_prices.py 抓的最新價)
  2. 產生「本月快照」JSON,存到 snapshots/{YYYY-MM-DD}.json
  3. 對比上月快照 → 計算 MoM 變化、最佳/最差個股
  4. 用 yfinance 抓近 30 日資料計算 Sharpe / 波動 / Alpha
  5. 組合 Markdown 報告 → Telegram 推播
  6. (由 workflow 後續 git commit 把 snapshot push 回 repo)
"""

import os
import sys
import json
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook

# 重用既有模組
import alerts as alerts_mod
from update_prices import TICKER_MAP, get_usd_twd

_TPE = ZoneInfo('Asia/Taipei')
SNAPSHOT_DIR = Path('snapshots')
XLSX_DEFAULT = '持股整合.xlsx'


# ============================================================
# 工具
# ============================================================
def _resolve_yf_symbol(code: str, is_us: bool) -> str:
    if is_us:
        return code
    if code in TICKER_MAP:
        _, yf_fb = TICKER_MAP[code]
        if yf_fb:
            return yf_fb
    return f'{code}.TW'


def _esc_md(s: str) -> str:
    """Telegram Markdown V1 不需太多 escape;但 ( ) _ * 仍可能影響格式。
    簡單把會造成解析問題的字元做替換。"""
    if not s:
        return ''
    s = str(s)
    # 替換 underscore (Telegram Markdown V1 會把 _xxx_ 視為 italic)
    return s.replace('_', '\\_')


# ============================================================
# 載入持股
# ============================================================
def load_holdings(xlsx_path: str) -> tuple[pd.DataFrame, float]:
    """從 xlsx 載入持股清單與 USDTWD 匯率。"""
    wb = load_workbook(xlsx_path, data_only=True)
    if '持股健檢策略' not in wb.sheetnames:
        raise ValueError(f'xlsx 缺少「持股健檢策略」分頁,現有: {wb.sheetnames}')

    ws = wb['持股健檢策略']
    rows = []
    for r in range(16, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        s = str(name).strip()
        if not s or '合計' in s or s in ('美股', '台股') or s.startswith(('USD', 'JPY')):
            continue
        code = alerts_mod._extract_code(s)
        # 欄位對應 (與 streamlit_app.py load_data 一致):
        #   A=name  B=category  C=shares  D=avg_cost  F=price
        category = ws.cell(r, 2).value or '其他'
        shares   = ws.cell(r, 3).value
        avg_cost = ws.cell(r, 4).value
        price    = ws.cell(r, 6).value
        if not isinstance(shares, (int, float)) or shares <= 0:
            continue
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        is_us = bool(code) and code[0].isalpha()
        rows.append({
            'code':     code,
            'name':     s,
            'shares':   float(shares),
            'avg_cost': float(avg_cost or 0),
            'price':    float(price),
            'category': str(category),
            'is_us':    is_us,
        })
    df = pd.DataFrame(rows)
    usd_twd = get_usd_twd(wb)
    return df, usd_twd


# ============================================================
# 快照處理
# ============================================================
def take_snapshot(df: pd.DataFrame, usd_twd: float) -> dict:
    """產生當前組合 snapshot dict (可序列化為 JSON)。"""
    df = df.copy()
    df['mv_twd']   = df.apply(lambda r: r['price']    * r['shares'] * (usd_twd if r['is_us'] else 1.0), axis=1)
    df['cost_twd'] = df.apply(lambda r: r['avg_cost'] * r['shares'] * (usd_twd if r['is_us'] else 1.0), axis=1)
    df['pnl_twd']  = df['mv_twd'] - df['cost_twd']

    total_mv   = float(df['mv_twd'].sum())
    total_cost = float(df['cost_twd'].sum())
    total_pnl  = total_mv - total_cost

    cat_breakdown = {str(k): float(v) for k, v in df.groupby('category')['mv_twd'].sum().items()}

    stocks = []
    for _, r in df.iterrows():
        stocks.append({
            'code':     r['code'],
            'name':     r['name'],
            'shares':   float(r['shares']),
            'price':    float(r['price']),
            'mv_twd':   float(r['mv_twd']),
            'pnl_twd':  float(r['pnl_twd']),
            'pnl_pct':  float((r['pnl_twd'] / max(r['cost_twd'], 1)) * 100),
            'is_us':    bool(r['is_us']),
            'category': r['category'],
        })

    return {
        'snapshot_date':     datetime.now(_TPE).strftime('%Y-%m-%d'),
        'usd_twd':           float(usd_twd),
        'total_mv_twd':      total_mv,
        'total_cost_twd':    total_cost,
        'total_pnl_twd':     total_pnl,
        'total_pnl_pct':     (total_pnl / max(total_cost, 1)) * 100,
        'n_stocks':          len(df),
        'category_breakdown': cat_breakdown,
        'stocks':            stocks,
    }


def load_previous_snapshot() -> dict | None:
    """載入「最新但不是今天」的 snapshot JSON。"""
    if not SNAPSHOT_DIR.exists():
        return None
    today = datetime.now(_TPE).strftime('%Y-%m-%d')
    files = sorted(SNAPSHOT_DIR.glob('*.json'))
    for f in reversed(files):
        if f.stem == today:
            continue
        try:
            with f.open(encoding='utf-8') as fp:
                return json.load(fp)
        except Exception:
            continue
    return None


def save_snapshot(snapshot: dict):
    SNAPSHOT_DIR.mkdir(exist_ok=True)
    fn = SNAPSHOT_DIR / f'{snapshot["snapshot_date"]}.json'
    with fn.open('w', encoding='utf-8') as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    print(f'📸 已儲存快照: {fn}')


# ============================================================
# 月度變化
# ============================================================
def compute_monthly_changes(curr: dict, prev: dict | None) -> dict:
    if not prev:
        return {'has_prev': False}

    mv_change = curr['total_mv_twd'] - prev['total_mv_twd']
    mv_pct    = (mv_change / max(prev['total_mv_twd'], 1)) * 100

    # Per-stock 漲跌幅 (依 code 對應)
    prev_stocks = {s['code']: s for s in prev.get('stocks', [])}
    movers = []
    for s in curr['stocks']:
        p = prev_stocks.get(s['code'])
        if not p:
            continue
        if p['price'] <= 0:
            continue
        price_pct = (s['price'] / p['price'] - 1) * 100
        movers.append({
            'name':       s['name'],
            'code':       s['code'],
            'price_pct':  price_pct,
            'mv_delta':   s['mv_twd'] - p['mv_twd'],
        })

    movers.sort(key=lambda x: x['price_pct'], reverse=True)
    best = movers[:3]
    worst = movers[-3:] if len(movers) >= 3 else movers[::-1]
    worst.reverse()   # 由跌幅最大開始

    return {
        'has_prev':  True,
        'prev_date': prev['snapshot_date'],
        'mv_change': mv_change,
        'mv_pct':    mv_pct,
        'best':      best,
        'worst':     worst,
    }


# ============================================================
# 30 日風險指標
# ============================================================
def fetch_30day_metrics(df: pd.DataFrame, usd_twd: float) -> dict:
    """用 yfinance 抓近 2 個月歷史,計算 30 日報酬與風險指標。"""
    try:
        import yfinance as yf
    except ImportError:
        return {}

    syms = {}
    for _, r in df.iterrows():
        sym = _resolve_yf_symbol(r['code'], r['is_us'])
        syms[sym] = (float(r['shares']), bool(r['is_us']))

    closes = {}
    for sym in syms:
        try:
            h = yf.Ticker(sym).history(period='2mo', auto_adjust=True)
            if h.empty or 'Close' not in h.columns:
                continue
            s = h['Close']
            if s.index.tz is not None:
                s.index = s.index.tz_localize(None)
            closes[sym] = s
        except Exception:
            continue

    if not closes:
        return {}

    # Benchmark
    bench = None
    try:
        h = yf.Ticker('0050.TW').history(period='2mo', auto_adjust=True)
        if not h.empty and 'Close' in h.columns:
            s = h['Close']
            if s.index.tz is not None:
                s.index = s.index.tz_localize(None)
            bench = s
    except Exception:
        pass

    # Union of dates
    idx_set = set()
    for s in closes.values():
        idx_set.update(s.index)
    idx_sorted = sorted(idx_set)
    if len(idx_sorted) < 5:
        return {}
    # 只取近 30 個交易日
    last_30 = idx_sorted[-30:]

    # 組合每日市值
    pf_vals = []
    for dt in last_30:
        total = 0.0
        n_have = 0
        for sym, (shares, is_us) in syms.items():
            s = closes.get(sym)
            if s is None:
                continue
            avail = s[s.index <= dt]
            if avail.empty:
                continue
            p = float(avail.iloc[-1])
            total += shares * p * (usd_twd if is_us else 1.0)
            n_have += 1
        if total > 0 and n_have / max(len(syms), 1) >= 0.8:
            pf_vals.append((dt, total))

    if len(pf_vals) < 5:
        return {}

    pf = pd.Series({dt: v for dt, v in pf_vals}).sort_index()
    daily_ret = pf.pct_change().dropna()
    if daily_ret.empty:
        return {}

    vol_ann   = float(daily_ret.std() * (252 ** 0.5) * 100)
    ret_ann   = float(daily_ret.mean() * 252 * 100)
    sharpe    = (ret_ann - 1.5) / vol_ann if vol_ann > 0 else 0.0
    cummax    = pf.cummax()
    mdd       = float(((pf - cummax) / cummax * 100).min())
    month_ret = float((pf.iloc[-1] / pf.iloc[0] - 1) * 100) if pf.iloc[0] > 0 else 0.0

    bench_ret = None
    if bench is not None and not bench.empty:
        bench_30 = bench.reindex(pf.index, method='ffill').dropna()
        if len(bench_30) >= 2 and bench_30.iloc[0] > 0:
            bench_ret = float((bench_30.iloc[-1] / bench_30.iloc[0] - 1) * 100)

    return {
        'month_return':       month_ret,
        'bench_return':       bench_ret,
        'annualized_vol':     vol_ann,
        'sharpe':             sharpe,
        'mdd_30d':            mdd,
    }


# ============================================================
# Markdown 組裝
# ============================================================
def build_report(curr: dict, changes: dict, metrics: dict) -> str:
    now = datetime.now(_TPE)
    lines = []
    lines.append(f'📅 *{now.strftime("%Y/%m")} 月度組合報告*')
    lines.append(f'_{now.strftime("%Y-%m-%d %H:%M TPE")}_')
    lines.append('')

    # 總覽
    lines.append('💰 *總覽*')
    lines.append(f'• 總市值: *{curr["total_mv_twd"]:,.0f}* TWD')
    lines.append(f'• 總損益: *{curr["total_pnl_twd"]:+,.0f}* ({curr["total_pnl_pct"]:+.2f}%)')
    lines.append(f'• 持股: {curr["n_stocks"]} 檔　USDTWD: {curr["usd_twd"]:.2f}')
    lines.append('')

    # 月度變化
    if changes.get('has_prev'):
        lines.append(f'📊 *月度變化* (vs {changes["prev_date"]})')
        arrow = '📈' if changes['mv_change'] >= 0 else '📉'
        lines.append(f'• 市值變動: *{changes["mv_change"]:+,.0f}* TWD ({changes["mv_pct"]:+.2f}%) {arrow}')
        lines.append('')
        if changes.get('best'):
            lines.append('🟢 *最佳 3 檔*')
            for s in changes['best']:
                lines.append(f'  • {_esc_md(s["name"])} *{s["price_pct"]:+.2f}%*')
            lines.append('')
        if changes.get('worst'):
            lines.append('🔴 *最差 3 檔*')
            for s in changes['worst']:
                lines.append(f'  • {_esc_md(s["name"])} *{s["price_pct"]:+.2f}%*')
            lines.append('')
    else:
        lines.append('📊 _首次快照,下月起會比對 MoM 變化_')
        lines.append('')

    # 風險指標
    if metrics:
        lines.append('🛡️ *風險指標 (近 30 個交易日)*')
        lines.append(f'• 月報酬: *{metrics["month_return"]:+.2f}%*')
        if metrics.get('bench_return') is not None:
            alpha = metrics['month_return'] - metrics['bench_return']
            arrow = '🟢' if alpha >= 0 else '🔴'
            lines.append(f'• 0050 月報酬: {metrics["bench_return"]:+.2f}%')
            lines.append(f'• 超額報酬: *{alpha:+.2f}%* {arrow}')
        lines.append(f'• 年化波動: {metrics["annualized_vol"]:.2f}%')
        lines.append(f'• Sharpe: *{metrics["sharpe"]:.2f}*')
        lines.append(f'• 30 日 MDD: {metrics["mdd_30d"]:.2f}%')
        lines.append('')

    # 類別配置 Top 6
    if curr.get('category_breakdown'):
        lines.append('📊 *類別配置*')
        cats = sorted(curr['category_breakdown'].items(), key=lambda x: -x[1])[:6]
        for cat, mv in cats:
            pct = mv / curr['total_mv_twd'] * 100 if curr['total_mv_twd'] > 0 else 0
            lines.append(f'  • {_esc_md(cat)}: {pct:.1f}% ({mv:,.0f})')
        lines.append('')

    lines.append('_🤖 由 GitHub Actions 自動產生_')
    return '\n'.join(lines)


# ============================================================
# Telegram 推播
# ============================================================
def send_telegram(message: str) -> bool:
    import requests
    token   = os.getenv('TG_BOT_TOKEN')
    chat_id = os.getenv('TG_CHAT_ID')
    if not token or not chat_id:
        print('⚠ TG_BOT_TOKEN / TG_CHAT_ID 未設定,跳過推播')
        print('--- 報告內容 ---')
        print(message)
        print('---')
        return False
    url = f'https://api.telegram.org/bot{token}/sendMessage'
    try:
        r = requests.post(url, json={
            'chat_id':                chat_id,
            'text':                   message,
            'parse_mode':             'Markdown',
            'disable_web_page_preview': True,
        }, timeout=20)
        if r.status_code == 200:
            print(f'📨 月報推送成功 ({len(message)} chars)')
            return True
        print(f'❌ Telegram API {r.status_code}: {r.text[:200]}')
        return False
    except Exception as e:
        print(f'❌ 推播 exception: {str(e)[:120]}')
        return False


# ============================================================
# main
# ============================================================
def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else XLSX_DEFAULT
    if not os.path.exists(xlsx):
        print(f'❌ 找不到 {xlsx}')
        sys.exit(1)

    print(f'🔄 開始產生月報 ({datetime.now(_TPE).strftime("%Y-%m-%d %H:%M TPE")})')

    df, usd_twd = load_holdings(xlsx)
    print(f'✓ 載入 {len(df)} 檔持股,USDTWD = {usd_twd:.2f}')

    curr = take_snapshot(df, usd_twd)
    print(f'✓ 當前總市值 {curr["total_mv_twd"]:,.0f} TWD')

    prev = load_previous_snapshot()
    if prev:
        print(f'✓ 比對上月快照: {prev["snapshot_date"]} ({prev["total_mv_twd"]:,.0f})')
    else:
        print('ℹ️  無前期快照,首次執行')

    changes = compute_monthly_changes(curr, prev)
    metrics = fetch_30day_metrics(df, usd_twd)

    report = build_report(curr, changes, metrics)
    print('\n' + '=' * 60)
    print(report)
    print('=' * 60 + '\n')

    send_telegram(report)
    save_snapshot(curr)

    print('✅ 月報完成')


if __name__ == '__main__':
    main()
